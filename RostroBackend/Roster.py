from pandas import to_datetime
from openpyxl import workbook, worksheet
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Fill
import os, re
import datetime
import xlrd
import openpyxl
from icalendar import Calendar, Event
from icalendar.cal import Todo

# RosterType
#     -> EDRoster
#     -> MAPURosteMAPURoster
# Interface

def military_string_to_datetimetime(time: str) -> datetime.time:
    if len(time) != 4:
        raise ValueError(time)
    if time[2] == '0' and time[0] == '0':
        time = datetime.time(int(time[1]),int(time[3]))
    elif time[2] == '0' and time[0] != '0':
        time = datetime.time(int(time[:2]),int(time[3]))
    elif time[2] != '0' and time[0] == '0':
        time = datetime.time(int(time[1]),int(time[-2:]))
    else:
        time = datetime.time(int(time[:2]),int(time[-2:]))
    return time

def extract_shift_from_string(string:str):
    if not isinstance(string, str):
        return string
    if string == 'OFF' or string == 'ON CALL':
        x = [string]
        return x
    x = re.findall('[0-9]{4}|[0-9]{1,2}[\:][0-9]{2}', string)
    if x == []:
        return None
    new_x = []
    for item in x:
        item = military_string_to_datetimetime(item)
        new_x.append(item)
    return new_x

def open_xls_as_xlsx(filename):
    book_xls = xlrd.open_workbook(filename)
    book_xlsx = openpyxl.Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                cell = sheet_xls.cell_value(row,col)
                if isinstance(cell,float):
                    cell = xlrd.xldate_as_datetime(cell, book_xls.datemode)
                    today = datetime.date.today()
                    if today.year -1 < cell.year < today.year+1:
                        sheet_xlsx.cell(row = row+1 , column = col+1).value = cell
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
    new_filename = filename.split('.')[0] + '.xlsx'
    book_xlsx.save(new_filename)
    # os.remove(filename)
    return new_filename

class Roster():
    def __init__(self, username: str, rosterpath: str) -> None:
        self.rosterpath = rosterpath
        self.rostername = os.path.basename(rosterpath)
        self.username = username
        self.wb = self.load_roster_from_file()[1]
        self.df = self.load_roster_from_file()[0]
        self.legend = self.get_legend()
        self.fullshifts = self.get_fullshifts()
        self.shiftcount = self.get_shiftcount()
    
    def load_roster_from_file(self): # add return type
        print(self.rosterpath)
        basename = os.path.basename(self.rosterpath)
        print(basename.split('.'))
        if basename.split('.')[1] == 'xls':
            self.rosterpath = open_xls_as_xlsx(self.rosterpath)
        wb = load_workbook(self.rosterpath)
        df = wb.worksheets[0]
        return (df,wb)
    
    def get_dates(self) -> list[tuple]:
        '''returns any datetime or pd.Timestamp along with its index as a list of lists'''
        dates = []
        for rowindex, row in enumerate(self.df.iter_rows()):
            for index, workcell in enumerate(row):
                if isinstance(workcell, openpyxl.cell.cell.Cell) and workcell.is_date and workcell.value != None:
                    # if int(workcell.value.strftime('%W')) + 12 < int(datetime.date.today().strftime('%W')):
                    if int(workcell.value.strftime('%W')) + 26 < int(datetime.date.today().strftime('%W')):
                        workcell.value = workcell.value.replace(year = workcell.value.year+1)
                    date = workcell.value
                    dates.append((index,rowindex,date))
        return dates
    
    def get_legend(self) -> None:
        legend_index = None
        for row in self.df.iter_rows():
            for cell in row:
                if cell.value == 'Legend':
                    legend_index = cell
                    break
            else:
                continue
            break
        if legend_index == None:
            return None
        legend = []
        for col in self.df.iter_cols(min_row=legend_index.row, min_col = legend_index.column, max_col= legend_index.column):
            for cell in col:
                if cell.value != None:
                    value = cell.value
                    print(cell.fill.start_color.theme, cell.value)
                    if isinstance(cell.fill.start_color.theme,int):
                        legend.append((value, cell.fill.start_color.theme))
                    if cell.fill.start_color.rgb and not isinstance(cell.fill.start_color.theme,int):
                        if value == 'Red':
                            value = 'Acute Medical'
                        if value == 'Yellow':
                            value = 'Acute Trauma'
                        if value == 'Green':
                            value = 'Ambulatory Care'
                        legend.append((value, cell.fill.start_color.rgb))
        print(legend)
        return legend

    def get_shifts(self) -> list[tuple]:
            if self.nameindex == None:
                print(f'No shifts in week of {self.df.title}')
                return None
            shiftextract = None
            shiftdata = []
            number_of_rows = 1
            for col in self.df.iter_cols(min_col = self.nameindex[0], max_col= self.nameindex[0]):
                user_below_bool = False
                user_above_bool = False
                print(type(col[self.nameindex[1]-1-4].value))
                while user_above_bool != True and user_below_bool != True:
                    above_value = col[self.nameindex[1]-1-number_of_rows].value
                    try:
                        below_value = col[self.nameindex[1]-1+number_of_rows].value
                    except IndexError:
                        below_value = None
                    if isinstance(above_value, str) and re.match('[+-]?([0-9]*[.])?[0-9]+',above_value):
                        above_value = int(above_value)
                    if isinstance(below_value, str) and re.match('[+-]?([0-9]*[.])?[0-9]+',below_value):
                        below_value = int(below_value)
                    user_above_bool = isinstance(above_value,str)
                    user_below_bool = isinstance(below_value,str)
                    if user_above_bool == False and user_below_bool == False:
                        number_of_rows = number_of_rows + 1
            # if number_of_rows == 1:
            #     for row in self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1):
            #         shiftextract = row
            #         break
            # else:
            #     shiftextract = self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1)
            for date in self.dates:
                for row in self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1):
                    if isinstance(row[date[0]].value, str):
                        shift_time_list = extract_shift_from_string(row[date[0]].value)
                    if shift_time_list == None:
                        continue
                    shift_start = shift_time_list[0]
                    if len(shift_time_list) == 2:
                        shift_end = shift_time_list[1]
                    elif len(shift_time_list) == 1:
                        shift_end = extract_shift_from_string(row[date[0]+1].value)
                        if shift_end != None:
                            shift_end = shift_end[0]
                    startdate = date[2]
                    enddate = date[2]
                    print(shift_end)
                    if isinstance(shift_start, datetime.date) and isinstance (shift_end, datetime.date) and shift_start > shift_end:
                        enddate = enddate + datetime.timedelta(days=1)
                shiftdata.append((startdate, shift_start, enddate, shift_end, row[date[0]]))
            print(shiftdata)
            print(len(shiftdata))
            return shiftdata

    def format_shifts(self):
        if self.shifts == None:
            return None
        new_shifts = []
        for shift in self.shifts:
            if shift[1] == 'OFF' or shift[3] == None:
                continue           
            description = None
            if self.legend != None:
                for key in self.legend:
                    if shift[4].fill.start_color.theme== key[1]:
                        description = key[0]
                    elif shift[4].fill.start_color.rgb== key[1]:
                        description = key[0]
            formatted_shift = (shift[0], shift[1], shift[2], shift[3], description)
            if shift[1] == 'ON CALL':
                formatted_shift = (shift[0], datetime.time(0,0), shift[0], datetime.time(23,59), description)
            new_shifts.append(formatted_shift)
        return new_shifts

    def rotate_worksheets(self):
        sheetnames = self.wb.sheetnames
        for i, sheetname in enumerate(sheetnames):
            if self.df.title == sheetname and i+1 != len(sheetnames):
                self.df = self.wb[sheetnames[i+1]]
                break
            elif self.df.title == sheetname and i == len(sheetnames):
                break
        return

    def get_fullshifts(self) -> None:
        fullshifts = []
        for sheet in self.wb.sheetnames:
            self.initialise()
            self.shifts = self.format_shifts()
            if self.shifts != None:
                for shift in self.shifts:
                    fullshifts.append(shift)
            self.rotate_worksheets()
        return fullshifts

    def initialise(self):
        self.nameindex = self.get_indexes()
        self.dates = self.get_dates()
        self.shifts = self.get_shifts()

    def get_indexes(self) -> list[int]:
        usernames = re.split('[^a-zA-Z]', self.username)
        usernames = list(filter(None, usernames))
        for rows in self.df.iter_rows():
            for cell in rows:
                if isinstance(cell.value,str) and all(re.search(username, cell.value, flags=re.IGNORECASE) for username in usernames):
                    index = (cell.col_idx, cell.row)
                    print(index, self.username)
                    return index
    
    def get_shiftcount(self):
        i = len(self.fullshifts)
        return i
    
    def create_ical(self) -> None:
        if self.fullshifts == None:
            print('EDRoster fullshifts is empty','\t', self.__str__ )
            return None
        cal = Calendar()
        caldir = os.path.dirname(self.rosterpath)
        # username = self.username.replace(' ','_')
        usernames = re.split('[^a-zA-Z]', self.username)
        usernames = list(filter(None, usernames))
        username = '-'.join(usernames)
        calname = f'{username}-Rostro.ics'
        calpath = os.path.join(caldir,calname)
        for shift in self.fullshifts:
            event = Event()
            datetimestart = datetime.datetime.combine(shift[0],shift[1])
            datetimeend = datetime.datetime.combine(shift[2], shift[3])
            if shift[4] != None:
                event.add('summary', shift[4])
            else:
                event.add('summary', 'Shift')
            event.add('dtstart', datetimestart)
            event.add('dtend', datetimeend)
            cal.add_component(event)
        calstring = cal.to_ical()
        if os.path.isfile(calpath):
            print('overwriting file')
            writemethod = 'wb'
        else:
            writemethod = 'xb'
        print(f'found {self.shiftcount} shifts')
        with open(calpath, writemethod) as f:
            f.write(calstring)
        return calpath

