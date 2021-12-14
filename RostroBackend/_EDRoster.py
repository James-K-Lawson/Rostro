from re import match
import datetime as datetime
from icalendar import Calendar, Event
from icalendar.cal import Todo
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Fill
# from RostroBackend.Roster import *
from Roster import *
import os
import re

class EDRoster(Roster):
    def __init__(self, username: str, rosterpath: str) -> None:
        super().__init__(username, rosterpath)
        print(self.username, self.rostername)
        self.wb = self.load_roster_from_file()[1]
        self.df = self.load_roster_from_file()[0]
        self.legend = self.get_legend()
        self.fullshifts = self.get_fullshifts()
        self.shiftcount = self.get_shiftcount()
    
    def __str__(self):
        string = (
            "Roster object"
            f"{self}.wb = {self.wb} \t"
            f"{self}.df = {self.df}  \t"
            f"{self}.legend = {self.legend}  \t"
            f"{self}.fullshifts = {self.fullshifts}  \t"
            f"{self}.shiftcount = {self.shiftcount} \t"
        )
        return string
    
    def initialise(self):
        self.nameindex = self.get_indexes()
        self.dates = self.get_dates()
        self.shifts = self.get_shifts()

    def get_indexes(self) -> list[int]:
        usernames = re.split('[^a-zA-Z]', self.username)
        usernames = list(filter(None, usernames))
        for rows in self.df.iter_rows():
            for cell in rows:
                if isinstance(cell.value,str) and all(re.search(username, cell.value, flags=re.IGNORECASE) for username in usernames):
                    index = (cell.col_idx, cell.row)
                    print(index, self.username)
                    return index

    def get_shifts(self) -> list[tuple]:
        if self.nameindex == None:
            print(f'No shifts in week of {self.df.title}')
            return None
        shiftextract = None
        shiftdata = []
        number_of_rows = 1
        for col in self.df.iter_cols(min_col = self.nameindex[0], max_col= self.nameindex[0]):
            user_below_bool = False
            user_above_bool = False
            print(type(col[self.nameindex[1]-1-4].value))
            while user_above_bool != True and user_below_bool != True:
                above_value = col[self.nameindex[1]-1-number_of_rows].value
                try:
                    below_value = col[self.nameindex[1]-1+number_of_rows].value
                except IndexError:
                    below_value = None
                if isinstance(above_value, str) and re.match('[+-]?([0-9]*[.])?[0-9]+',above_value):
                    above_value = int(above_value)
                if isinstance(below_value, str) and re.match('[+-]?([0-9]*[.])?[0-9]+',below_value):
                    below_value = int(below_value)
                user_above_bool = isinstance(above_value,str)
                user_below_bool = isinstance(below_value,str)
                if user_above_bool == False and user_below_bool == False:
                    number_of_rows = number_of_rows + 1
        # if number_of_rows == 1:
        #     for row in self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1):
        #         shiftextract = row
        #         break
        # else:
        #     shiftextract = self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1)
        for date in self.dates:
            for row in self.df.iter_rows(min_row=self.nameindex[1], max_row= self.nameindex[1]+number_of_rows-1):
                if isinstance(row[date[0]].value, str):
                    shift_time_list = extract_shift_from_string(row[date[0]].value)
                if shift_time_list == None:
                    continue
                shift_start = shift_time_list[0]
                if len(shift_time_list) == 2:
                    shift_end = shift_time_list[1]
                elif len(shift_time_list) == 1:
                    shift_end = extract_shift_from_string(row[date[0]+1].value)
                    if shift_end != None:
                        shift_end = shift_end[0]
                startdate = date[2]
                enddate = date[2]
                print(shift_end)
                if isinstance(shift_start, datetime.date) and isinstance (shift_end, datetime.date) and shift_start > shift_end:
                    enddate = enddate + datetime.timedelta(days=1)
            shiftdata.append((startdate, shift_start, enddate, shift_end, row[date[0]]))
        print(shiftdata)
        print(len(shiftdata))
        return shiftdata
    
    def format_shifts(self):
        if self.shifts == None:
            return None
        new_shifts = []
        for shift in self.shifts:
            if shift[1] == 'OFF' or shift[3] == None:
                continue           
            description = None
            if self.legend != None:
                for key in self.legend:
                    if shift[4].fill.start_color.theme== key[1]:
                        description = key[0]
                    elif shift[4].fill.start_color.rgb== key[1]:
                        description = key[0]
            formatted_shift = (shift[0], shift[1], shift[2], shift[3], description)
            if shift[1] == 'ON CALL':
                formatted_shift = (shift[0], datetime.time(0,0), shift[0], datetime.time(23,59), description)
            new_shifts.append(formatted_shift)
        return new_shifts

    def rotate_worksheets(self):
        sheetnames = self.wb.sheetnames
        for i, sheetname in enumerate(sheetnames):
            if self.df.title == sheetname and i+1 != len(sheetnames):
                self.df = self.wb[sheetnames[i+1]]
                break
            elif self.df.title == sheetname and i == len(sheetnames):
                break
        return

    def get_fullshifts(self) -> None:
        fullshifts = []
        for sheet in self.wb.sheetnames:
            self.initialise()
            self.shifts = self.format_shifts()
            if self.shifts != None:
                for shift in self.shifts:
                    fullshifts.append(shift)
            self.rotate_worksheets()
        return fullshifts
    
    def get_shiftcount(self):
        i = len(self.fullshifts)
        return i

# def _add_dates():
#         dates = []
#         ed = EDRoster('Term 5 2021 House Officer.xlsx', 'XU, Jack')
#         workcell = ed.df.cell(row = 3, column = 3)
#         workcell.value = datetime.date(2021,1,1)
#         if isinstance(workcell, openpyxl.cell.cell.Cell) and workcell.is_date and workcell.value != None:
#             print('accessed1')
#             if int(workcell.value.strftime('%W')) + 12 < int(datetime.date.today().strftime('%W')):
#                 print('accessed')
#                 workcell.value = workcell.value.replace(year = workcell.value.year+1)
#             print(workcell.value)
#             date = workcell.value
#             dates.append(date)

if __name__ == "__main__":
    ed = EDRoster('James Lawson', 'MAPURoster.xls')
    ed.create_ical()