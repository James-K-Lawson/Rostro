from re import match
import datetime as datetime
from icalendar import Calendar, Event
from icalendar.cal import Todo
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Fill
from RostroBackend.Roster import *
import os


def military_string_to_datetimetime(time: str) -> datetime.time:
    if len(time) != 4:
        raise ValueError(time)
    if time[2] == '0' and time[0] == '0':
        time = datetime.time(int(time[1]),int(time[3]))
    elif time[2] == '0' and time[0] != '0':
        time = datetime.time(int(time[:2]),int(time[3]))
    elif time[2] != '0' and time[0] == '0':
        time = datetime.time(int(time[1]),int(time[-2:]))
    else:
        time = datetime.time(int(time[:2]),int(time[-2:]))
    return time

class EDRoster(Roster):
    def __init__(self, username: str, rosterpath: str) -> None:
        super().__init__(username, rosterpath)
        print(self.username, self.rostername)
        self.wb = self.load_roster_from_file()[1]
        self.df = self.load_roster_from_file()[0]
        self.legend = self.get_legend()
        self.fullshifts = self.get_fullshifts()
        self.shiftcount = self.get_shiftcount()
    
    def load_roster_from_file(self): # add return type
        print(self.rosterpath)
        wb = load_workbook(self.rosterpath)
        df = wb.worksheets[0]
        return (df,wb)
    
    def initialise(self):
        self.nameindex = self.get_indexes()
        self.dates = self.get_dates()
        self.shifts = self.get_shifts()

    def print_roster(self) -> None:
        print(type(self.df))
        print(self.df.columns)
        rows = self.df.iter_rows()
        for i in rows:
            for cell in i:
                print(cell.value)

    def get_indexes(self) -> list[int]:

        for rows in self.df.iter_rows():
            for cell in rows:
                if cell.value == self.username:
                    index = (cell.col_idx, cell.row)
                    print(index, self.username)
                    return index


    def get_dates(self) -> list[tuple]:
        '''returns any datetime or pd.Timestamp along with its index as a list of lists'''
        dates = []
        for rowindex, row in enumerate(self.df.iter_rows()):
            for index, workcell in enumerate(row):
                if isinstance(workcell, openpyxl.cell.cell.Cell) and workcell.is_date and workcell.value != None:
                    if int(workcell.value.strftime('%W')) + 12 < int(datetime.date.today().strftime('%W')):
                        workcell.value = workcell.value.replace(year = workcell.value.year+1)
                    date = workcell.value
                    dates.append((index,rowindex,date))
        return dates
    
    def get_legend(self) -> None:
        legend_index = None
        for row in self.df.iter_rows():
            for cell in row:
                if cell.value == 'Legend':
                    legend_index = cell
                    break
            else:
                continue
            break
        legend = []
        for col in self.df.iter_cols(min_row=legend_index.row, min_col = legend_index.column, max_col= legend_index.column):
            for cell in col:
                if cell.value != None:
                    value = cell.value
                    print(cell.fill.start_color.theme, cell.value)
                    if isinstance(cell.fill.start_color.theme,int):
                        legend.append((value, cell.fill.start_color.theme))
                    if cell.fill.start_color.rgb and not isinstance(cell.fill.start_color.theme,int):
                        if value == 'Red':
                            value = 'Acute Medical'
                        if value == 'Yellow':
                            value = 'Acute Trauma'
                        if value == 'Green':
                            value = 'Ambulatory Care'
                        legend.append((value, cell.fill.start_color.rgb))
        print(legend)
        return legend

    def get_shifts(self) -> list[tuple]:
        if self.nameindex == None:
            print(f'No shifts in week of {self.df.title}')
            return None
        shiftextract = None
        shiftdata = []
        for row in self.df.iter_rows(min_row=self.nameindex[1]):
            shiftextract = row
            break

        for date in self.dates:
            shiftdata.append((date[2], shiftextract[date[0]].value, shiftextract[date[0]], self.nameindex[1]))
        return shiftdata
    
    def format_shifts(self):
        if self.shifts == None:
            return None
        new_shifts = []
        for shift in self.shifts:
            if shift[1] == 'OFF':
                continue           
            description = None
            for key in self.legend:
                if shift[2].fill.start_color.theme== key[1]:
                    description = key[0]
                elif shift[2].fill.start_color.rgb== key[1]:
                    description = key[0]
            formatted_time = (shift[1].split('-'))
            formatted_time = list(formatted_time)
            formatted_shift = None
            if len(formatted_time) == 2:
                for i, time in enumerate(formatted_time):
                    time = military_string_to_datetimetime(time)
                    formatted_time[i] = time         
                formatted_shift = (shift[0],formatted_time[0],formatted_time[1], shift[2], shift[3], description)
            elif shift[1] == 'ON CALL':
                formatted_shift = (shift[0], datetime.time(0,0), datetime.time(23,59), shift[2], shift[3], description)
            new_shifts.append(formatted_shift)
        return new_shifts

    def rotate_worksheets(self):
        sheetnames = self.wb.sheetnames
        for i, sheetname in enumerate(sheetnames):
            if self.df.title == sheetname and i+1 != len(sheetnames):
                self.df = self.wb[sheetnames[i+1]]
                break
            elif self.df.title == sheetname and i == len(sheetnames):
                break
        return

    def get_fullshifts(self) -> None:
        fullshifts = []
        for sheet in self.wb.sheetnames:
            self.initialise()
            self.shifts = self.format_shifts()
            if self.shifts != None:
                for shift in self.shifts:
                    fullshifts.append(shift)
            self.rotate_worksheets()
        return fullshifts
    
    def get_shiftcount(self):
        i = len(self.fullshifts)
        return i

    def create_ical(self) -> None:
        cal = Calendar()
        caldir = os.path.dirname(self.rosterpath)
        calname = f'{self.username} Rostro.ics'
        calpath = os.path.join(caldir,calname)
        for shift in self.fullshifts:
            event = Event()
            datetimestart = datetime.datetime.combine(shift[0],shift[1])
            datetimeend = datetime.datetime.combine(shift[0], shift[2])
            event.add('summary', shift[5])
            event.add('dtstart', datetimestart)
            event.add('dtend', datetimeend)
            cal.add_component(event)
        calstring = cal.to_ical()
        if os.path.isfile(calpath):
            writemethod = 'wb'
        else:
            writemethod = 'xb'
        with open(calpath, writemethod) as f:
            f.write(calstring)
        return calpath

# def _add_dates():
#         dates = []
#         ed = EDRoster('Term 5 2021 House Officer.xlsx', 'XU, Jack')
#         workcell = ed.df.cell(row = 3, column = 3)
#         workcell.value = datetime.date(2021,1,1)
#         if isinstance(workcell, openpyxl.cell.cell.Cell) and workcell.is_date and workcell.value != None:
#             print('accessed1')
#             if int(workcell.value.strftime('%W')) + 12 < int(datetime.date.today().strftime('%W')):
#                 print('accessed')
#                 workcell.value = workcell.value.replace(year = workcell.value.year+1)
#             print(workcell.value)
#             date = workcell.value
#             dates.append(date)

if __name__ == '__main__':
    ed = EDRoster('Term_5_2021_Intern.xlsx','LAWSON, James')
    ed.run_rostro()